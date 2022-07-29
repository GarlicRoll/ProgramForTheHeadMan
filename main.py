import time

import telebot
import openpyxl
import os

def init_bot(): # инициализация бота
    print("Bot_init")

    '''
    with open("bot_token.txt", "r") as file:
        telegram_token = file.readline()
        global bot
        bot = telebot.TeleBot(telegram_token)
    '''

    telegram_token = os.environ.get('TOKEN')
    global bot
    bot = telebot.TeleBot(telegram_token)


def mark_a_person(message): # отметить человека в excel-таблице
    group = "Group n1307"
    try:
        try:
            book = openpyxl.load_workbook(filename="attendance.xlsx")
        except FileNotFoundError:
            book = openpyxl.Workbook()
        try:
            sheet = book[group]
        except KeyError:
            book.create_sheet(group)
            sheet = book[group]
        x, y = 2, 1
        while sheet.cell(row=y, column=x).value is not None:
            if sheet.cell(row=y, column=1).value == message.from_user.first_name:
                while sheet.cell(row=y, column=x).value is not None:
                    x += 1
            else:
                y += 1
        sheet.cell(row=y, column=1, value=message.from_user.first_name)
        sheet.cell(row=y, column=x, value=time.ctime(time.time()))
        book.save("attendance.xlsx")
    except PermissionError:
        bot.send_message(message.chat.id, 'Ведутся технические работы, вы не отмечены')

if __name__ == "__main__":
    bot = None
    init_bot()
    print("Start")

@bot.message_handler(commands=['test', 'start', 'location'])
def statistic_command(message): # запрос местоположения
    Geo = telebot.types.ReplyKeyboardMarkup(True, True)
    button_geo = telebot.types.KeyboardButton("Отметиться на паре", request_location=True)
    Geo.add(button_geo) # добавление кнопки
    # создаём объект закрытия кастомной клавиатуры
    Geo_close = telebot.types.ReplyKeyboardRemove()
    bot.send_message(message.chat.id, 'Поделитесь местоположением', reply_markup=Geo )

@bot.message_handler(commands=['statistics', 'students', 'info'])
def statistic_command(message): # запрос местоположения
    group = "Group n1307"
    try:
        book = openpyxl.load_workbook(filename="attendance.xlsx")
    except FileNotFoundError:
        bot.send_message(message.chat.id, 'Данные о посещаемости не найдены')
    try:
        sheet = book[group]
        x, y = 1, 1
        text = "Данные:\n"
        while sheet.cell(row=y, column=x).value is not None:
            while sheet.cell(row=y, column=x).value is not None:
                if x != 1:
                    text += "\t"
                text += str(sheet.cell(row=y, column=x).value) + "\n"
                x += 1
            y += 1
            x = 1
            text += "\n"
        bot.send_message(message.chat.id, text)
    except KeyError:
        bot.send_message(message.chat.id, 'Данные о группе "' + group + '" не найдены')


@bot.message_handler(content_types=['location'])
def handle_loc(message):
    longitude = message.location.longitude
    latitude = message.location.latitude
    # создание файла
    '''
    message_buffer_id = 100
    file_buffer_id = 110
    with open("data.txt", "r+") as file:
        #file.write(str(longitude) + " " + str(latitude))
        file.write('test text')
        bot.edit_message_media(chat_id=message.chat.id, message_id=message_buffer_id, media=file_buffer_id)
        #print(bot.send_document(message.chat.id, file).id)
    #print(message.id)
    '''
    epsilon = 0.005
    # проверка местоположения
    if (30.2967 - epsilon < round(longitude, 4) < 30.2967 + epsilon) and (59.9717 - epsilon < round(latitude, 4) < 59.9717 + epsilon): #59.97168732680863, 30.296766628835186
        text = "Молодец, " + str(message.from_user.first_name) + ", ходи на пары в 6 корпус дальше!"
        mark_a_person(message)
    elif (30.3227 - epsilon < round(longitude, 4) < 30.3227 + epsilon) and (59.9723 - epsilon < round(latitude, 4) < 59.9723 + epsilon): #59.97225610894957, 30.322730700399283
        text = "Молодец, " + str(message.from_user.first_name) + ", ходи на пары дальше!"
        mark_a_person(message)
    else:
        text = "ААА, " + str(message.from_user.first_name) + "!!! ТЫ НЕ В ЛЭТИ!!!"
        # mark_a_person(message) # Для тестирования
    bot.send_message(message.chat.id, text)

bot.polling()